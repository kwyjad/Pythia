# Pythia
# Copyright (c) 2025 Kevin Wyjad
# Licensed under the Pythia Non-Commercial Public License v1.0.
# See the LICENSE file in the project root for details.

"""Dartmouth Flood Observatory archive — a CALIBRATION CROSS-CHECK ONLY.

The DFO Global Active Archive of Large Flood Events is an independent,
manually curated record of large floods reaching back to 1985. GDACS —
which is what the machine detects floods with — only stabilises around
2010, which is why ``backcast.flood`` starts there.

That leaves an obvious question the backcast cannot answer about itself:
*is the flood occurrence rate the machine derives from its own detector
plausible?* Comparing the machine against the years it already resolved
would confirm nothing. So this module compares the machine's occurrence
rates against DFO's, over the years strictly BEFORE the backcast era. A
country whose DFO-era rate is far above the machine's rate is a country
where flood detection is probably missing events — a finding for a human,
recorded in the acceptance report.

**What this module may never do.** DFO is not a trigger and not a
resolution value. It writes to exactly one table, ``haz_raw_dfo``, and
returns comparison rows to the report. It imports no writer for
``haz_triggers``, ``haz_impact_candidates`` or ``haz_resolutions``, and a
test asserts that those tables are untouched by everything in here. The
reason is the same one that keeps GDACS exposure out of the ladder: an
archive assembled from press reporting is a different quantity from a
resolved people-affected figure, and a source admitted "just for
calibration" is one refactor away from being admitted as an answer.

**Endpoint caveat.** ``dfo.url`` in the rulebook could not be reached from
the environment this was built in. A wrong URL or a changed sheet shape
fails LOUD and SAFE: the fetch reports unavailable, the report's
cross-check section says so, and nothing else in the machine is affected —
because nothing else reads this source. The fix is a URL in YAML and at
worst a column name in :data:`_COLUMN_ALIASES`.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import logging
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Callable, Mapping

import requests

from resolver.hazard_resolution.rulebook import Rulebook
from resolver.hazard_resolution.sources import (
    FetchOutcome,
    RawRecord,
    load_raw_records,
    parse_date,
    store_raw_records,
    utcnow_iso,
)

if TYPE_CHECKING:  # pragma: no cover - typing only
    import duckdb

LOG = logging.getLogger(__name__)

SOURCE = "dfo"

#: The repo hazard code this archive speaks about. DFO records floods and
#: nothing else, so the cross-check is flood-only by construction.
HAZARD = "FL"

#: Injectable transport seam for tests: (url, timeout) -> raw bytes.
GetFn = Callable[[str, float], bytes]

#: DFO has published the archive under several column spellings across
#: releases (and the xlsx and csv exports do not always agree). Each field
#: lists the spellings the reader accepts, matched case-insensitively after
#: stripping spaces and underscores. Being tolerant here is deliberate: a
#: renamed column should cost a line in this table, not a debugging session.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("id", "register#", "registernumber", "glidenumber"),
    "country": ("country", "maincountry", "countryname"),
    "other_country": ("othercountry", "othercountries", "additionalcountries"),
    "began": ("began", "begandate", "startdate", "start"),
    "ended": ("ended", "endeddate", "enddate", "end"),
    "dead": ("dead", "deaths", "fatalities"),
    "displaced": ("displaced", "displacedpersons", "affected"),
    "severity": ("severity", "severity*", "severityclass"),
    "main_cause": ("maincause", "cause"),
}


def _norm_header(value: Any) -> str:
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def _map_columns(headers: list[Any]) -> dict[str, int]:
    """Map our field names onto the sheet's column indices."""

    normalised = [_norm_header(h) for h in headers]
    out: dict[str, int] = {}
    for field_name, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                out[field_name] = normalised.index(alias)
                break
    return out


def _http_get(url: str, timeout: float) -> bytes:
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    return response.content


def _rows_from_csv(blob: bytes) -> list[list[Any]]:
    text = blob.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(str(c).strip() for c in row)]


def _rows_from_xlsx(blob: bytes) -> list[list[Any]]:
    # openpyxl is already a first-class dependency (the repo's exporters use
    # it), so the xlsx path costs no new install.
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
    finally:
        workbook.close()


def _iso3_for(name: Any) -> str | None:
    """Resolve a DFO country label to ISO3, or None.

    Uses the repo's shared resolver (and therefore its alias table), so a
    name form fixed for one connector is fixed for this one too.
    """

    text = str(name or "").strip()
    if not text:
        return None
    from resolver.ingestion.utils.iso_normalize import to_iso3

    return to_iso3(text)


@dataclass
class DfoEvent:
    """One archive entry, reduced to what the cross-check reads."""

    record_id: str
    iso3: str
    began: dt.date
    ended: dt.date | None = None
    detail: dict[str, Any] = field(default_factory=dict)

    @property
    def year(self) -> int:
        return self.began.year

    @property
    def month(self) -> int:
        return self.began.month


def parse_archive(
    blob: bytes, fmt: str
) -> tuple[list[DfoEvent], dict[str, Any]]:
    """Parse the archive into events plus a parse report. Never raises on data.

    The report counts what was dropped and why. A silent drop and "the
    archive was thin that decade" are indistinguishable downstream, and this
    module's whole job is to say whether a rate looks too low.
    """

    rows = _rows_from_csv(blob) if fmt == "csv" else _rows_from_xlsx(blob)
    report: dict[str, Any] = {
        "rows": max(0, len(rows) - 1),
        "parsed": 0,
        "dropped_no_iso3": 0,
        "dropped_no_date": 0,
        "unresolved_country_samples": [],
    }
    if not rows:
        report["error"] = "archive is empty"
        return [], report

    columns = _map_columns(rows[0])
    missing = [f for f in ("country", "began") if f not in columns]
    if missing:
        report["error"] = (
            f"archive is missing required column(s) {missing}; headers seen: "
            f"{[str(h) for h in rows[0]][:20]}"
        )
        return [], report
    report["columns"] = {k: str(rows[0][v]) for k, v in columns.items()}

    def cell(row: list[Any], name: str) -> Any:
        index = columns.get(name)
        return row[index] if index is not None and index < len(row) else None

    events: list[DfoEvent] = []
    unresolved: set[str] = set()
    for index, row in enumerate(rows[1:], start=1):
        began = parse_date(cell(row, "began"))
        if began is None:
            report["dropped_no_date"] += 1
            continue
        # The archive names a main country and sometimes a spillover list.
        # Every named country counts as having had the flood: a flood that
        # crossed a border occurred in both, and the cross-check asks about
        # occurrence, not about attribution of a figure.
        names = [cell(row, "country")]
        others = str(cell(row, "other_country") or "").strip()
        if others:
            names.extend(part for part in others.replace(";", ",").split(",") if part.strip())

        record_id = str(cell(row, "id") or f"row-{index}").strip() or f"row-{index}"
        matched = False
        for position, name in enumerate(names):
            iso3 = _iso3_for(name)
            if iso3 is None:
                if str(name or "").strip():
                    unresolved.add(str(name).strip())
                continue
            matched = True
            events.append(
                DfoEvent(
                    record_id=f"{record_id}-{iso3}" if position else record_id,
                    iso3=iso3,
                    began=began,
                    ended=parse_date(cell(row, "ended")),
                    detail={
                        "dfo_id": record_id,
                        "country_label": str(name).strip(),
                        "dead": cell(row, "dead"),
                        "displaced": cell(row, "displaced"),
                        "severity": cell(row, "severity"),
                        "main_cause": cell(row, "main_cause"),
                        "spillover": bool(position),
                    },
                )
            )
        if not matched:
            report["dropped_no_iso3"] += 1

    report["parsed"] = len(events)
    report["unresolved_country_samples"] = sorted(unresolved)[:10]
    report["dropped_unresolved_names"] = len(unresolved)
    return events, report


def fetch_dfo(
    con: "duckdb.DuckDBPyConnection",
    rulebook: Rulebook,
    *,
    get: GetFn | None = None,
) -> FetchOutcome:
    """Download and cache the DFO archive. Never raises.

    Like every connector in the machine, an unreachable source is reported
    as UNAVAILABLE rather than as an empty archive — here that means the
    report says the cross-check could not run, which is a different
    statement from "the machine's rates agree with DFO".
    """

    if not bool(rulebook.get("dfo.enabled")):
        return FetchOutcome(
            source=SOURCE, ok=False, error="dfo.enabled is false — cross-check not consulted"
        )

    url = str(rulebook.get("dfo.url"))
    fmt = str(rulebook.get("dfo.format"))
    timeout = float(rulebook.get("dfo.request_timeout_sec"))
    getter = get or _http_get

    try:
        blob = getter(url, timeout)
    except Exception as exc:  # noqa: BLE001 - a source outage is data, not a crash
        LOG.warning("[dfo] archive fetch failed (%s): %s", url, exc)
        return FetchOutcome(source=SOURCE, ok=False, source_urls=[url], error=str(exc))

    try:
        events, report = parse_archive(blob, fmt)
    except Exception as exc:  # noqa: BLE001 - a changed sheet shape is data too
        LOG.warning("[dfo] archive parse failed (%s): %s", url, exc)
        return FetchOutcome(source=SOURCE, ok=False, source_urls=[url], error=str(exc))

    if report.get("error"):
        LOG.warning("[dfo] archive unusable: %s", report["error"])
        return FetchOutcome(
            source=SOURCE, ok=False, source_urls=[url], error=str(report["error"]),
            detail=report,
        )
    # A parse that resolves nothing has changed shape. Refusing it is the
    # same rule the drought-indicator layer applies: a feed where 0 of N
    # records resolve is a broken feed, not a world without floods.
    if not events:
        LOG.warning("[dfo] archive resolved 0 of %s rows to ISO3 — refusing", report["rows"])
        return FetchOutcome(
            source=SOURCE,
            ok=False,
            source_urls=[url],
            error=f"0 of {report['rows']} archive rows resolved to a country",
            detail=report,
        )

    counts = store_raw_records(
        con,
        SOURCE,
        [
            RawRecord(
                record_id=event.record_id,
                payload={
                    "iso3": event.iso3,
                    "began": event.began.isoformat(),
                    "ended": event.ended.isoformat() if event.ended else None,
                    **event.detail,
                },
                iso3=event.iso3,
                ym=f"{event.year:04d}-{event.month:02d}",
                hazard=HAZARD,
                source_url=url,
            )
            for event in events
        ],
    )
    LOG.info(
        "[dfo] cached %d archive events (%d new rows) from %s",
        len(events), counts["inserted"], url,
    )
    return FetchOutcome(
        source=SOURCE,
        ok=True,
        records=len(events),
        inserted=counts["inserted"],
        source_urls=[url],
        detail={**report, "retrieved_at": utcnow_iso()},
    )


def load_events(
    con: "duckdb.DuckDBPyConnection", *, before_year: int | None = None
) -> list[DfoEvent]:
    """Cached archive events, optionally limited to years before ``before_year``."""

    events: list[DfoEvent] = []
    for payload in load_raw_records(con, SOURCE, hazard=HAZARD):
        began = parse_date(payload.get("began"))
        iso3 = str(payload.get("iso3") or "").upper()
        if began is None or len(iso3) != 3:
            continue
        if before_year is not None and began.year >= int(before_year):
            continue
        events.append(
            DfoEvent(
                record_id=str(payload.get("_record_id") or ""),
                iso3=iso3,
                began=began,
                ended=parse_date(payload.get("ended")),
                detail={k: v for k, v in payload.items() if not k.startswith("_")},
            )
        )
    return events


def occurrence_rates(
    events: list[DfoEvent],
) -> dict[tuple[str, int], dict[str, Any]]:
    """DFO occurrence rate per (iso3, calendar month). Pure function.

    Same shape as the machine's own occurrence base rate — the share of
    observed years in which at least one flood began in that calendar month
    — so the two are directly comparable.

    The denominator is the archive's own span, not each country's: a
    country that appears in DFO only in wet years would otherwise show a
    rate of 1.0 for every month it appears in, which is an artefact of
    counting only the years it was seen.
    """

    if not events:
        return {}
    years = [event.year for event in events]
    first_year, last_year = min(years), max(years)
    n_years = last_year - first_year + 1

    seen: dict[tuple[str, int], set[int]] = {}
    for event in events:
        seen.setdefault((event.iso3, event.month), set()).add(event.year)

    return {
        key: {
            "p_occurrence": len(observed) / n_years,
            "n_years_with_event": len(observed),
            "n_years": n_years,
            "window": f"{first_year}-{last_year}",
        }
        for key, observed in seen.items()
    }


@dataclass
class CrossCheckRow:
    """One country's machine-era rate beside its DFO-era rate."""

    iso3: str
    machine_p: float
    dfo_p: float
    machine_window: str
    dfo_window: str
    n_months_compared: int
    ratio: float | None
    diverges: bool


@dataclass
class CrossCheck:
    """The whole comparison, plus why it could not run if it could not."""

    available: bool
    reason: str | None = None
    rows: list[CrossCheckRow] = field(default_factory=list)
    dfo_window: str | None = None
    machine_window: str | None = None
    divergence_factor: float | None = None
    n_events: int = 0

    @property
    def diverging(self) -> list[CrossCheckRow]:
        return [row for row in self.rows if row.diverges]


def cross_check(
    con: "duckdb.DuckDBPyConnection", rulebook: Rulebook
) -> CrossCheck:
    """Compare the machine's flood occurrence rates against DFO's.

    Reads ``haz_base_rates_occurrence`` (which the backcast produced) and
    the cached archive. Both sides are annualised per country — the mean of
    its twelve monthly rates — because a month-by-month comparison of two
    sparse records is noise, while "how often does this country flood in a
    year" is a quantity both archives can answer.

    Writes nothing anywhere.
    """

    if not bool(rulebook.get("dfo.enabled")):
        return CrossCheck(available=False, reason="dfo.enabled is false")

    end_year = int(rulebook.get("dfo.cross_check_end_year"))
    min_years = int(rulebook.get("dfo.min_years_for_comparison"))
    factor = float(rulebook.get("dfo.divergence_factor"))

    events = load_events(con, before_year=end_year)
    if not events:
        return CrossCheck(
            available=False,
            reason=(
                f"no cached DFO events before {end_year} — run the fetch, or the "
                "archive endpoint is unreachable (see rulebook dfo.url)"
            ),
        )

    dfo = occurrence_rates(events)
    if not dfo:
        return CrossCheck(available=False, reason="DFO events produced no rates")
    dfo_window = next(iter(dfo.values()))["window"]
    if int(dfo_window.split("-")[1]) - int(dfo_window.split("-")[0]) + 1 < min_years:
        return CrossCheck(
            available=False,
            reason=(
                f"DFO window {dfo_window} is shorter than "
                f"dfo.min_years_for_comparison={min_years}"
            ),
        )

    machine_rows = con.execute(
        """
        SELECT iso3, calendar_month, p_occurrence, source_window
        FROM haz_base_rates_occurrence
        WHERE hazard = ?
        ORDER BY iso3, calendar_month
        """,
        [HAZARD],
    ).fetchall()
    if not machine_rows:
        return CrossCheck(
            available=False,
            reason=(
                "haz_base_rates_occurrence has no flood rows — compute the base "
                "rates from the backcast before cross-checking them"
            ),
            n_events=len(events),
            dfo_window=dfo_window,
        )

    machine: dict[str, list[float]] = {}
    machine_window = None
    for iso3, month, p, window in machine_rows:
        machine.setdefault(str(iso3), []).append(float(p))
        machine_window = machine_window or window

    rows: list[CrossCheckRow] = []
    for iso3, monthly in sorted(machine.items()):
        # Annualised expected months-with-a-flood on each side. Summing the
        # monthly rates (rather than averaging) keeps the number readable as
        # "flood-months per year", which is what a human wants to compare.
        machine_p = sum(monthly)
        dfo_p = sum(
            dfo.get((iso3, month), {}).get("p_occurrence", 0.0) for month in range(1, 13)
        )
        if machine_p <= 0 and dfo_p <= 0:
            continue  # both archives say this country does not flood
        ratio = (dfo_p / machine_p) if machine_p > 0 else None
        diverges = (
            ratio is None  # DFO records floods the machine's era never sees
            or ratio > factor
            or ratio < 1.0 / factor
        )
        rows.append(
            CrossCheckRow(
                iso3=iso3,
                machine_p=machine_p,
                dfo_p=dfo_p,
                machine_window=str(machine_window or ""),
                dfo_window=dfo_window,
                n_months_compared=len(monthly),
                ratio=ratio,
                diverges=diverges,
            )
        )

    LOG.info(
        "[dfo] cross-check: %d countries compared (%s vs %s), %d diverge by > %.1fx",
        len(rows), machine_window, dfo_window,
        sum(1 for r in rows if r.diverges), factor,
    )
    return CrossCheck(
        available=True,
        rows=rows,
        dfo_window=dfo_window,
        machine_window=str(machine_window or ""),
        divergence_factor=factor,
        n_events=len(events),
    )
